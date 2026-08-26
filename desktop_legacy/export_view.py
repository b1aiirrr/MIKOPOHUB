
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from database import get_connection


# ======================================================
# EXPORT WINDOW
# ======================================================

class ExportView:

    def __init__(self, parent):

        self.parent = parent

        self.window = tk.Toplevel(parent)

        self.window.title(
            "MikopoHub - Download Reports"
        )

        self.window.geometry(
            "600x500"
        )

        self.window.resizable(
            False,
            False
        )

        self.build_interface()

    # ==================================================
    # BUILD INTERFACE
    # ==================================================

    def build_interface(self):

        container = tk.Frame(
            self.window,
            padx=30,
            pady=25
        )

        container.pack(
            fill="both",
            expand=True
        )

        # ----------------------------------------------
        # TITLE
        # ----------------------------------------------

        tk.Label(
            container,
            text="📥 Download Reports",
            font=("Arial", 22, "bold")
        ).pack(
            pady=(0, 10)
        )

        tk.Label(
            container,
            text=(
                "Select the information you want to "
                "download and print."
            ),
            font=("Arial", 10)
        ).pack(
            pady=(0, 25)
        )

        # ----------------------------------------------
        # REPORT TYPE
        # ----------------------------------------------

        tk.Label(
            container,
            text="Select Report",
            font=("Arial", 11, "bold")
        ).pack(
            anchor="w",
            pady=(10, 5)
        )

        self.report_var = tk.StringVar()

        self.report_combo = ttk.Combobox(
            container,
            textvariable=self.report_var,
            state="readonly",
            values=[
                "Dashboard Summary",
                "Borrowers Report",
                "Loans Report",
                "Payments Report",
                "Push Forward History",
                "Collateral Report"
            ],
            font=("Arial", 11)
        )

        self.report_combo.pack(
            fill="x",
            pady=(0, 20)
        )

        # ----------------------------------------------
        # INFORMATION
        # ----------------------------------------------

        info_frame = tk.Frame(
            container,
            relief="solid",
            borderwidth=1,
            padx=15,
            pady=15
        )

        info_frame.pack(
            fill="x",
            pady=10
        )

        tk.Label(
            info_frame,
            text=(
                "The report will be saved as an Excel "
                "file (.xlsx).\n\n"
                "You can open it using Microsoft Excel "
                "or another spreadsheet application "
                "and print it."
            ),
            justify="left",
            font=("Arial", 10)
        ).pack()

        # ----------------------------------------------
        # DOWNLOAD BUTTON
        # ----------------------------------------------

        tk.Button(
            container,
            text="📥 DOWNLOAD REPORT",
            font=("Arial", 11, "bold"),
            command=self.download_report
        ).pack(
            fill="x",
            pady=(25, 10)
        )

        # ----------------------------------------------
        # CLOSE
        # ----------------------------------------------

        tk.Button(
            container,
            text="Close",
            command=self.window.destroy
        ).pack(
            fill="x",
            pady=5
        )

    # ==================================================
    # DOWNLOAD REPORT
    # ==================================================

    def download_report(self):

        report = self.report_var.get()

        if not report:

            messagebox.showwarning(
                "No Report Selected",
                "Please select a report first."
            )

            return

        if report == "Dashboard Summary":

            self.export_dashboard_summary()

        elif report == "Borrowers Report":

            self.export_borrowers()

        elif report == "Loans Report":

            self.export_loans()

        elif report == "Payments Report":

            self.export_payments()

        elif report == "Push Forward History":

            self.export_push_forward_history()

        elif report == "Collateral Report":

            self.export_collateral()

    # ==================================================
    # CREATE WORKBOOK
    # ==================================================

    def create_workbook(
        self,
        title,
        headers,
        rows
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = title[:31]

        # ----------------------------------------------
        # REPORT TITLE
        # ----------------------------------------------

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(headers)
        )

        title_cell = worksheet.cell(
            row=1,
            column=1
        )

        title_cell.value = (
            f"MIKOPOHUB - {title.upper()}"
        )

        title_cell.font = Font(
            bold=True,
            size=16
        )

        title_cell.alignment = Alignment(
            horizontal="center"
        )

        # ----------------------------------------------
        # REPORT DATE
        # ----------------------------------------------

        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(headers)
        )

        date_cell = worksheet.cell(
            row=2,
            column=1
        )

        date_cell.value = (
            f"Generated on: {date.today()}"
        )

        date_cell.alignment = Alignment(
            horizontal="center"
        )

        # ----------------------------------------------
        # HEADERS
        # ----------------------------------------------

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        header_row = 4

        for column_number, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number
            )

            cell.value = header

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

            cell.border = thin_border

        # ----------------------------------------------
        # DATA
        # ----------------------------------------------

        current_row = header_row + 1

        for row in rows:

            for column_number, value in enumerate(
                row,
                start=1
            ):

                cell = worksheet.cell(
                    row=current_row,
                    column=column_number
                )

                cell.value = value

                cell.border = thin_border

            current_row += 1

        # ----------------------------------------------
        # COLUMN WIDTH
        # ----------------------------------------------

        for column in worksheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    if cell.value:

                        cell_length = len(
                            str(cell.value)
                        )

                        if cell_length > max_length:

                            max_length = cell_length

                except Exception:

                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40
            )

        return workbook

    # ==================================================
    # SAVE WORKBOOK
    # ==================================================

    def save_workbook(
        self,
        workbook,
        default_name
    ):

        file_path = filedialog.asksaveasfilename(

            title="Save MikopoHub Report",

            defaultextension=".xlsx",

            initialfile=default_name,

            filetypes=[
                (
                    "Excel Workbook",
                    "*.xlsx"
                ),
                (
                    "All Files",
                    "*.*"
                )
            ]
        )

        if not file_path:

            return

        try:

            workbook.save(
                file_path
            )

            messagebox.showinfo(
                "Download Complete",
                "Report downloaded successfully!\n\n"
                f"Saved to:\n{file_path}"
            )

        except Exception as error:

            messagebox.showerror(
                "Download Error",
                f"Unable to save report.\n\n"
                f"{error}"
            )

    # ==================================================
    # DASHBOARD SUMMARY
    # ==================================================

    def export_dashboard_summary(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                "SELECT COUNT(*) FROM borrowers"
            )

            borrowers = cursor.fetchone()[0] or 0

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM loans
                WHERE status = 'ACTIVE'
                """
            )

            active_loans = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COALESCE(
                    SUM(principal),
                    0
                )
                FROM loans
                WHERE status != 'VOID'
                """
            )

            total_lent = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COALESCE(
                    SUM(principal_portion),
                    0
                )
                FROM payments
                """
            )

            principal_paid = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COALESCE(
                    SUM(principal),
                    0
                )
                FROM loans
                WHERE status = 'ACTIVE'
                """
            )

            active_principal = (
                cursor.fetchone()[0] or 0
            )

            outstanding = max(
                0,
                active_principal - principal_paid
            )

            cursor.execute(
                """
                SELECT COALESCE(
                    SUM(interest_portion),
                    0
                )
                FROM payments
                """
            )

            interest_collected = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COALESCE(
                    SUM(fee_amount),
                    0
                )
                FROM form_fees
                WHERE payment_status = 'PAID'
                """
            )

            form_fees = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM loans
                WHERE status = 'ACTIVE'
                AND due_date = date('now')
                """
            )

            due_today = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM loans
                WHERE status = 'ACTIVE'
                AND due_date < date('now')
                """
            )

            overdue = (
                cursor.fetchone()[0] or 0
            )

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM collateral
                WHERE status = 'HELD'
                """
            )

            collateral_held = (
                cursor.fetchone()[0] or 0
            )

            headers = [
                "Item",
                "Value"
            ]

            rows = [
                ["Total Borrowers", borrowers],
                ["Active Loans", active_loans],
                ["Total Lent (KSh)", total_lent],
                [
                    "Principal Outstanding (KSh)",
                    outstanding
                ],
                [
                    "Interest Collected (KSh)",
                    interest_collected
                ],
                [
                    "Form Fees Collected (KSh)",
                    form_fees
                ],
                ["Due Today", due_today],
                ["Overdue Loans", overdue],
                ["Collateral Held", collateral_held]
            ]

            workbook = self.create_workbook(
                "Dashboard Summary",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Dashboard_Summary.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to generate dashboard report.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()

    # ==================================================
    # BORROWERS REPORT
    # ==================================================

    def export_borrowers(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT
                    borrower_number,
                    full_name,
                    phone,
                    national_id,
                    address,
                    date_registered

                FROM borrowers

                ORDER BY full_name
                """
            )

            results = cursor.fetchall()

            headers = [
                "Borrower Number",
                "Full Name",
                "Phone",
                "National ID",
                "Address",
                "Date Registered"
            ]

            rows = []

            for borrower in results:

                rows.append(
                    [
                        borrower["borrower_number"],
                        borrower["full_name"],
                        borrower["phone"],
                        borrower["national_id"],
                        borrower["address"],
                        borrower["date_registered"]
                    ]
                )

            workbook = self.create_workbook(
                "Borrowers Report",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Borrowers_Report.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to export borrowers.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()

    # ==================================================
    # LOANS REPORT
    # ==================================================

    def export_loans(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT
                    loans.loan_number,
                    borrowers.full_name,
                    loans.principal,
                    loans.interest_rate,
                    loans.issue_date,
                    loans.due_date,
                    loans.status

                FROM loans

                JOIN borrowers
                    ON loans.borrower_id = borrowers.id

                ORDER BY loans.id DESC
                """
            )

            results = cursor.fetchall()

            headers = [
                "Loan Number",
                "Borrower",
                "Principal (KSh)",
                "Interest Rate (%)",
                "Date Borrowed",
                "Due Date",
                "Status"
            ]

            rows = []

            for loan in results:

                rows.append(
                    [
                        loan["loan_number"],
                        loan["full_name"],
                        loan["principal"],
                        loan["interest_rate"],
                        loan["issue_date"],
                        loan["due_date"],
                        loan["status"]
                    ]
                )

            workbook = self.create_workbook(
                "Loans Report",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Loans_Report.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to export loans.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()

    # ==================================================
    # PAYMENTS REPORT
    # ==================================================

    def export_payments(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT
                    loans.loan_number,
                    borrowers.full_name,
                    payments.amount,
                    payments.interest_portion,
                    payments.principal_portion,
                    payments.payment_date,
                    payments.payment_method

                FROM payments

                JOIN loans
                    ON payments.loan_id = loans.id

                JOIN borrowers
                    ON loans.borrower_id = borrowers.id

                ORDER BY payments.id DESC
                """
            )

            results = cursor.fetchall()

            headers = [
                "Loan Number",
                "Borrower",
                "Payment Amount (KSh)",
                "Interest Portion (KSh)",
                "Principal Portion (KSh)",
                "Payment Date",
                "Payment Method"
            ]

            rows = []

            for payment in results:

                rows.append(
                    [
                        payment["loan_number"],
                        payment["full_name"],
                        payment["amount"],
                        payment["interest_portion"],
                        payment["principal_portion"],
                        payment["payment_date"],
                        payment["payment_method"]
                    ]
                )

            workbook = self.create_workbook(
                "Payments Report",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Payments_Report.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to export payments.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()

    # ==================================================
    # PUSH FORWARD HISTORY
    # ==================================================

    def export_push_forward_history(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT
                    loans.loan_number,
                    borrowers.full_name,
                    push_forward_history.old_due_date,
                    push_forward_history.new_due_date,
                    push_forward_history.months_pushed,
                    push_forward_history.reason,
                    push_forward_history.created_at

                FROM push_forward_history

                JOIN loans
                    ON push_forward_history.loan_id = loans.id

                JOIN borrowers
                    ON loans.borrower_id = borrowers.id

                ORDER BY push_forward_history.id DESC
                """
            )

            results = cursor.fetchall()

            headers = [
                "Loan Number",
                "Borrower",
                "Old Due Date",
                "New Due Date",
                "Months Pushed",
                "Reason",
                "Date Created"
            ]

            rows = []

            for item in results:

                rows.append(
                    [
                        item["loan_number"],
                        item["full_name"],
                        item["old_due_date"],
                        item["new_due_date"],
                        item["months_pushed"],
                        item["reason"],
                        item["created_at"]
                    ]
                )

            workbook = self.create_workbook(
                "Push Forward History",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Push_Forward_History.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to export push forward history.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()

    # ==================================================
    # COLLATERAL REPORT
    # ==================================================

    def export_collateral(self):

        connection = None

        try:

            connection = get_connection()

            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT
                    collateral.collateral_number,
                    loans.loan_number,
                    borrowers.full_name,
                    collateral.security_type,
                    collateral.description,
                    collateral.estimated_value,
                    collateral.serial_number,
                    collateral.condition,
                    collateral.date_received,
                    collateral.status

                FROM collateral

                JOIN loans
                    ON collateral.loan_id = loans.id

                JOIN borrowers
                    ON loans.borrower_id = borrowers.id

                ORDER BY collateral.id DESC
                """
            )

            results = cursor.fetchall()

            headers = [
                "Collateral Number",
                "Loan Number",
                "Borrower",
                "Security Type",
                "Description",
                "Estimated Value (KSh)",
                "Serial Number",
                "Condition",
                "Date Received",
                "Status"
            ]

            rows = []

            for item in results:

                rows.append(
                    [
                        item["collateral_number"],
                        item["loan_number"],
                        item["full_name"],
                        item["security_type"],
                        item["description"],
                        item["estimated_value"],
                        item["serial_number"],
                        item["condition"],
                        item["date_received"],
                        item["status"]
                    ]
                )

            workbook = self.create_workbook(
                "Collateral Report",
                headers,
                rows
            )

            self.save_workbook(
                workbook,
                "MikopoHub_Collateral_Report.xlsx"
            )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                f"Unable to export collateral.\n\n"
                f"{error}"
            )

        finally:

            if connection:

                connection.close()